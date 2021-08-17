import datetime, os
import sqlite3
from pathlib import Path

import pyttsx3
from PyQt5 import QAxContainer
from PyQt5.QtWidgets import QApplication, QLabel, QDialog, QStackedWidget, QInputDialog, QLineEdit, QMessageBox, \
    QCommandLinkButton, QProgressBar, QWidget, QScrollArea
from PyQt5.QtGui import QPixmap, QIntValidator
from PyQt5.uic import loadUi
from PyQt5.QtCore import Qt, QDate, QTimer
import sys
import openpyxl as xl
import random
import smtplib
import ssl


class Window(QDialog):
    def __init__(self):
        super(Window, self).__init__()
        self.setWindowTitle("The Loveworld Children Curriculum App")
        self.setGeometry(100, 100, 1500, 700)
        loadUi("firstpage.ui", self)
        self.setGeometry(100, 100, 381, 311)
        self.label_3.setPixmap(QPixmap("C://Users//HP//OneDrive//Desktop//beginning.png"))
        self.teacher.clicked.connect(self.openteacher)
        self.child.clicked.connect(self.openchild)

    def openchild(self):
        window2.setCurrentIndex(1)

    def openteacher(self):
        window2.setCurrentIndex(2)

class TeacherPage(QDialog):
    def __init__(self):
        super(TeacherPage, self).__init__()
        loadUi("teacherpage.ui", self)
        self.setStyleSheet("background-image: url(photo_2021-08-09_21-06-21.jpg)")
        self.signin.clicked.connect(self.signindata)
        self.forgotmenu.clicked.connect(self.forgotstatus)
        self.back.clicked.connect(self.go_back)

    def go_back(self):
        window2.setCurrentIndex(0)
        # 0 1 2 3

    def forgotstatus(self):
        codec, done1 = QInputDialog.getText(self, "Email", "Please enter your email:")

        corr = False
        data = {}
        conn = sqlite3.connect("teachers.db")
        cursor = conn.execute("SELECT TeacherName, Teacherid, TeachersChurch, TeachersZone, Gender, Birthday, CertificateStatus, Country, PhoneNo, Email, Password, State from teachers")
        for row in cursor:
            if row[9] == codec:
                data = {'name': row[0], 'id': row[1], 'church': row[2], 'zone': row[3], 'gender': row[4], 'birthday': row[5], 'status-of-certificate': row[6], 'country': row[7], 'phone': row[8], 'email': row[9], 'password': row[10], 'state': row[11]}
                corr = True
                break
            else:
                pass
        if not corr:
            QMessageBox.about(self, 'Error', "Oops! It look like your email doesn't exist in our database.")
        else:
            try:
                self.secret_number = random.randint(100000, 999999)
                port = 465  # For SSL
                smtp_server = "smtp.gmail.com"
                sender_email = "orieozichi@gmail.com"  # Enter your address
                receiver_email = str(codec)  # Enter receiver address
                password = "growingbetter985"
                message = f"""\
                                                Subject: Verification\n

                                                Hi, {data['name']}. You clicked the option for Forgot Password.
                                                Verification code: {self.secret_number}
                                                Type this verification code in the GenChamps app"""

                context = ssl.create_default_context()
                with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
                    server.login(sender_email, password)
                    server.sendmail(sender_email, receiver_email, message)
                inputted, en = QInputDialog.getInt(self, 'Forgot Password', 'Check your mail and you will see a verification number\n\nType the verification number below:')
                if str(inputted) == str(self.secret_number):
                     msg = QMessageBox.question(self, "Your wish", "Do you wish to reset your password", QMessageBox.Yes | QMessageBox.No)
                     if msg == QMessageBox.No:
                         pass
                     else:
                         new_pass, m = QInputDialog.getText(self, "New Password", "Please enter your new password", QLineEdit.Password)
                         if m:
                             confirm, n = QInputDialog.getText(self, 'Confirm', 'Please type in again the password that you typed in again to confirm', QLineEdit.Password)
                             if n:
                                 if confirm == new_pass:
                                     QMessageBox.about(self, 'Saved!', "Your password has been changed")
                                     cursor.execute(f'Update teachers set Password = {confirm} where UserName = {data["name"]}')
                                     conn.commit()
                                     cursor.close()
                                 else:
                                     QMessageBox.about(self, 'Oops!', "You didn't type in the same password as you did before.\n\nBut don't worry, You can always use the 'Forgot Password' Option and we will get back to you.")
                     self.associate_teacher(data['name'], data['country'], data['state'], data['church'], data['zone'])
            except:
                QMessageBox.about(self, "Oops", "It looks like you do not have an internet connection")

    def signindata(self):
        conn = sqlite3.connect("teachers.db")
        cursor = conn.execute("SELECT TeacherName, Teacherid, TeachersChurch, TeachersZone, Gender, Birthday, CertificateStatus, Country, PhoneNo, Email, Password, State from teachers")
        chosen = False
        data = {}
        username = self.username.text()
        password = self.password.text()
        for row in cursor:
            if row[0] == username:
                data = {'name': row[0], 'id': row[1], 'church': row[2], 'zone': row[3], 'gender': row[4], 'birthday': row[5], 'status-of-certificate': row[6], 'country': row[7], 'phone': row[8], 'email': row[9], 'password': row[10], 'state': row[11]}
                chosen = True
                break
            else:
                pass
        if not chosen:
            self.usererror.setText("The username you typed below does not exist")
            self.passworderror.setText("")
        else:
            if data['password'] != password:
                self.usererror.setText("")
                self.passworderror.setText("The password you typed in is incorrect")
            else:
                self.associate_teacher(data['name'], data['country'], data['state'], data['church'], data['zone'])

    def associate_teacher(self, names, country, state, church, zone):
        window8.new_name = names
        window8.country = country
        window8.state = state
        window8.church = church
        window8.zone = zone

        window2.setCurrentIndex(6)
class RegisterChild(QDialog):
    def __init__(self, new_name, country, state, church, zone):
        super(RegisterChild, self).__init__()
        loadUi("registerpage.ui", self)
        self.new_name = new_name
        self.country = country
        self.state = state
        self.church = church
        self.zone = zone
        self.submitted.clicked.connect(self.register_2)
        self.phone.setValidator(QIntValidator())
        self.phone.textChanged.connect(self.collect)

    def collect(self):
        try:
            if str(self.phone.text()[0]) == "0":
                text = self.phone.text()
                text = text.replace('0', '')
                self.phone.setText(text)
        except:
            pass
    def compile_age(self, birth_date):
        listed = str(birth_date).split('/')
        print(listed)
        birth = QDate(int(listed[0]), int(listed[1]), int(listed[2]))

        birth_year = birth.year()
        birth_month = birth.month()
        birth_day = birth.day()

        current = QDate.currentDate()
        # getting year and month day of current day
        current_year = current.year()
        current_month = current.month()
        current_day = current.day()

        # coverting dates into date object
        birth_date = datetime.date(birth_year, birth_month, birth_day)
        current_date = datetime.date(current_year, current_month, current_day)

        # getting difference in both the dates
        difference = current_date - birth_date

        # getting days from the difference
        difference = difference.days
        years = difference / 365.2422
        years = round(years)
        return years

    def register_2(self):
        self.child_nam = self.child_name.text()
        self.birth_date = self.date_of_birth.date()
        self.parent_nam = self.parent_name.text()
        self.emails = self.email.text()
        self.phone_no = self.phone.text()

        self.birth_date_2 = self.compile_age(f'{self.birth_date.year()}/{self.birth_date.month()}/{self.birth_date.day()}')

        if self.child_nam == '':
            self.errors.setText("Please enter child name")
        else:
            if self.birth_date_2 > 13 or self.birth_date_2 < 2:
                self.errors.setText(f"Please the child you registered must \nbe between the ages 2-12 {self.birth_date}")
            else:
                if self.parent_nam == '':
                    self.errors.setText("Please the parent's name is important")
                else:
                    if self.emails == '':
                        self.errors.setText("Please enter the parent's email")
                    else:
                        if self.phone_no == '':
                            self.errors.setText("Please enter the parent's phone number.")
                        else:
                            first = random.choice(
                                ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q",
                                 "r", "s", "t", "u", "v", "w", "x", "y", "z"])
                            second = random.choice(
                                ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q",
                                 "R", "S", "T", "U", "V", "W", "X", "Y", "Z"])
                            third = random.choice(["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"])
                            fourth = random.choice(["$", "#", "%", "^", "*", "A", "z", "V", "c"])
                            fivth = random.choice(
                                ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q",
                                 "r", "s", "t", "u", "v", "w", "x", "y", "z"])
                            sixth = random.choice(
                                ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q",
                                 "R", "S", "T", "U", "V", "W", "X", "Y", "Z"])
                            seventh = random.choice(["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"])
                            eight = random.choice(["$", "#", "%", "^", "*", "A", "z", "V", "c"])
                            self.passcode = first + second + third + fourth + fivth + sixth + seventh + eight
                            #age = self.compile_age(f'{self.birth_date.year()}/{self.birth_date.month()}/{self.birth_date.day()}')
                            col = ''
                            if self.birth_date_2 <= 5 and self.birth_date_2 > 2:
                                col = 'Preschool'
                            elif self.birth_date_2 <= 9 and self.birth_date_2 > 5:
                                col = 'Elementary'
                            elif self.birth_date_2 <= 12 and self.birth_date_2 > 9:
                                col = 'Preteens'
                            data = {"name":self.child_nam, "password":self.passcode, "phone":self.phone_no, "email":self.emails, "teacher":self.new_name, 'classe':col, 'birth_date':f'{self.birth_date.year()}/{self.birth_date.month()}/{self.birth_date.day()}', 'acc_type':'CE Account', 'country':self.country, 'state':self.state, 'church':self.church, 'zone':self.zone, 'parent_name':self.parent_nam}
                            conn = sqlite3.connect("children.db")
                            cursor = conn.execute(f"""INSERT INTO children
                                                     (UserName, Password, PhoneNumber, Email, Teacher, Class, BirthDate, AccountType, Country, State, Church, Zone, ParentName, Coins)
                                                      VALUES
                                                      ('{data['name']}', '{data['password']}', '{data['phone']}', '{data['email']}', '{data['teacher']}', '{data['classe']}', '{data['birth_date']}', '{data['acc_type']}', '{data['country']}', '{data['state']}', '{data['church']}', '{data['zone']}', '{data['parent_name']}', 0)""")
                            conn.commit()
                            cursor.close()
class AssociateTeacher(QDialog):
    def __init__(self, new_name, country, state, church, zone):
        super(AssociateTeacher, self).__init__()
        loadUi("teacherpage2.ui", self)
        self.new_name = new_name
        self.country = country
        self.state = state
        self.church = church
        self.zone = zone
        self.label.setStyleSheet("background-image: url(welcometeacher.png);")
        self.label_2.setText(self.label_2.text().replace('[Teacher Name]', str(self.new_name)))
        self.registration.clicked.connect(self.register)
        self.child_curriculum.clicked.connect(self.show_curriculum)
        self.class_kids.clicked.connect(lambda: self.get_class_kids(new_name))

    def get_class_kids(self, name):
        w3 = ClassKids(name)
        window2.addWidget(w3)
        window2.setCurrentIndex(window2.currentIndex()+1)

    def get_teacher_children(self, name):
        try:
            conn = sqlite3.connect("children.db")
            cursor = conn.execute("SELECT UserName, Password, PhoneNumber, Email, Teacher, Class, BirthDate, AccountType, Country, State, Church, Zone, ParentName, Coins from children")
            data = {}
            conn = False
            for row in cursor:
                if row[4] == name:
                    child_name = row[0]
                    data[child_name] = {'name': child_name, 'phone': row[2],
                                        'email': row[3], 'class': row[5],
                                        'birth_date': row[6],
                                        'parent_name': row[12],
                                        'coins_made': row[13]}
            return data
        except Exception as e:
            print(e)

    def show_curriculum(self):
        w2 = Curriculums()
        window2.addWidget(w2)
        window2.setCurrentIndex(window2.currentIndex()+1)

    def register(self):
        w = RegisterChild(self.new_name, self.country, self.state, self.church, self.zone)
        window2.addWidget(w)
        window2.setCurrentIndex(window2.currentIndex()+1)

class ClassKids(QDialog):
    def __init__(self, teacher_name):
        super(ClassKids, self).__init__()
        loadUi("children.ui", self)
        self.teacher_name = teacher_name
        objects = self.get_teacher_children(self.teacher_name)
        text = 'Your Children:\n'
        for i in objects.values():
            text += f""""
            Child Name: {i['name']}
    Parent's Name: {i['parent_name']}
    Parent's Phone Number: {i['phone']}
    Parent's Email: {i['email']}
    Child Class: {i['class']}
    Child Birth Date: {i['birth_date']}
    Coins Made: {i['coins_made']}\n
    """
        self.label_2.setText(text)



    def get_teacher_children(self, name):
        try:
            conn = sqlite3.connect("children.db")
            cursor = conn.execute(
                "SELECT UserName, Password, PhoneNumber, Email, Teacher, Class, BirthDate, AccountType, Country, State, Church, Zone, ParentName, Coins from children")
            data = {}
            conn = False
            for row in cursor:
                if row[4] == name:
                    child_name = row[0]
                    data[child_name] = {'name': child_name, 'phone': row[2],
                                        'email': row[3], 'class': row[5],
                                        'birth_date': row[6],
                                        'parent_name': row[12],
                                        'coins_made': row[13]}
            return data
        except Exception as e:
            print(e)


class Curriculums(QDialog):
    def __init__(self):
        super(Curriculums, self).__init__()
        loadUi("curriculum_2.ui", self)
        self.lower_class.clicked.connect(self.open_lower)
        self.upper_class.clicked.connect(self.open_upper)

    def open_lower(self):
        wind2 = CurriculumPageLower()
        window2.addWidget(wind2)
        window2.setCurrentIndex(window2.currentIndex()+1)

    def open_upper(self):
        wind2 = CurriculumPageUpper()
        window2.addWidget(wind2)
        window2.setCurrentIndex(window2.currentIndex()+1)

class CurriculumPageLower(QDialog):
    def __init__(self):
        super(CurriculumPageLower, self).__init__()
        #loadUi("curriculum.ui", self)
        #self.label.setPixmap(QPixmap("C:\\Users\\HP\\PycharmProjects\\untitled\\GenChamps\\July\\0009"))
        self.WebBrowser = QAxContainer.QAxWidget(self)
        self.WebBrowser.setFocusPolicy(Qt.StrongFocus)
        self.WebBrowser.setControl("{8856F961-340A-11D0-A96B-00C04FD705A2}")
        self.WebBrowser.setFixedSize(354, 591)
        # convert system path to web path
        f = Path(f"{str(os.getcwd())}\lcbcforpresc_jul_2160df2e79b992b.pdf").as_uri()
        #f = Path(f'{str(os.getcwd())}\lcbcforupper_jul_2160df2ecaed84c.pdf').as_uri()
        # load object
        self.WebBrowser.dynamicCall('Navigate(const QString&)', f)
class CurriculumPageUpper(QDialog):
    def __init__(self):
        super(CurriculumPageUpper, self).__init__()
        #loadUi("curriculum.ui", self)
        #self.label.setPixmap(QPixmap("C:\\Users\\HP\\PycharmProjects\\untitled\\GenChamps\\July\\0009"))
        self.WebBrowser = QAxContainer.QAxWidget(self)
        self.WebBrowser.setFocusPolicy(Qt.StrongFocus)
        self.WebBrowser.setControl("{8856F961-340A-11D0-A96B-00C04FD705A2}")
        self.WebBrowser.setFixedSize(354, 591)
        # convert system path to web path
        f = Path(f'{str(os.getcwd())}\lcbcforupper_jul_2160df2ecaed84c.pdf').as_uri()
        # loadww object
        self.WebBrowser.dynamicCall('Navigate(const QString&)', f)
class ChildPage(QDialog):
    def __init__(self):
        super(ChildPage, self).__init__()
        loadUi("login.ui", self)
        self.signupp.clicked.connect(self.openchild2)
        self.signer.clicked.connect(self.check)
        self.forgot.clicked.connect(self.resolve)
        self.back.clicked.connect(self.go_back)

    def go_back(self):
        window2.setCurrentIndex(0)

    def resolve(self):
        email, entered = QInputDialog.getText(self, 'Forgot Password', 'Please enter your email:')
        conn = sqlite3.connect("children.db")
        cursor = conn.execute("SELECT UserName, Password, PhoneNumber, Email, Teacher, Class, BirthDate, AccountType, Country, State, Church, Zone, ParentName, Coins from children")
        chosen = False
        data = {}
        for row in cursor:
            if row[3] == email:
                data = {'name': row[0], 'password':row[1], 'phone':row[2], 'email':row[3], 'teacher':row[4], 'class':row[5], 'birth':row[6], 'acc_type':row[7], 'country':row[8], 'state':row[9], 'church':row[10], 'zone':row[11], 'parent-name': row[12], 'coins': row[13]}
                chosen = True
                break
            else:
                pass

        if chosen:
            try:
                self.secret_number = random.randint(100000, 999999)
                port = 465  # For SSL
                smtp_server = "smtp.gmail.com"
                sender_email = "orieozichi@gmail.com"  # Enter your address
                receiver_email = str(email)  # Enter receiver address
                password = "growingbetter985"
                message = f"""\
                                Subject: Verification\n
    
                                You clicked the option for Forgot Password.
                                Verification code: {self.secret_number}
                                Type this verification code in the GenChamps app"""

                context = ssl.create_default_context()
                with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
                    server.login(sender_email, password)
                    server.sendmail(sender_email, receiver_email, message)
                code, messed = QInputDialog.getInt(self, "Message", "Please check your mail and you will see a verification number\n\nType the verification number here:")
                if messed:
                    your_code = int(code)
                    if your_code != self.secret_number:
                        self.errors.setText("You entered a wrong number, please try again")
                    else:
                        new_pass, corr = QInputDialog.getText(self, "Password", "Please input your new password:", QLineEdit.Password)
                        if corr:
                            nexter, k = QInputDialog.getText(self, "Confirmation", "Please confirm your password", QLineEdit.Password)
                            if k:
                                if str(nexter) == str(new_pass):
                                    conn.execute(f"Update children set Password = {new_pass} where UserName = {data['name']}")
                                    conn.commit()
                                    cursor.close()
                                else:
                                    QMessageBox.about(self, "Error", "You didn't type in the same password but don't worry you can still use the 'Forgot Password' option to get into your account")

                        self.errors.setText("")
                        self.welcomeuser(data['name'], data['password'], data['phone'], data['email'], data['teacher'], data['class'], data['birth'], data['acc_type'], data['country'], data['state'], data['church'], data['zone'])
            except:
                QMessageBox.about(self, "Oops!", "It look like you do not have an internet connection")


        else:
            self.errors.setText("Your email does not exist in the app")

    def check(self):
        username = self.username.text()
        password = self.password.text()

        file = xl.load_workbook("children.xlsx")
        sheet = file['Sheet1']
        data = {}
        operation = 2
        correct = False
        users = []
        passwords = []
        point = 0
        conn = False

        if username == "":
            self.errors.setText("Please enter valid information")
        else:
            while not correct:
                if sheet["A"+str(operation)].value == None:
                    break
                else:
                    operation += 1

            for i2 in range(operation+1):
                users.append(sheet["A"+str(i2+1)].value)
                passwords.append(sheet["B"+str(i2+1)].value)

            for i in range(len(users)+1):
                if users[i-1] == username:
                    conn = True
                    point = i
                    break
                else:
                    continue

            if conn:
                for i in range(len(users)+1):
                    data[users[i-1]] = passwords[i-1]
                if data[username] != password:
                    self.errors.setText("Invalid Password")
                else:
                    self.child_info = {self.username.text():
                                           {'name':self.username.text(),
                                            'password':sheet['B'+str(point)].value,
                                            'phone':sheet['C'+str(point)].value,
                                            'email':sheet['D'+str(point)].value,
                                            'teacher':sheet['E'+str(point)].value,
                                            'classe':sheet['F'+str(point)].value,
                                            'birth_date':sheet['G'+str(point)].value,
                                            'acc_type':sheet['H'+str(point)].value,
                                            'country':sheet['I'+str(point)].value,
                                            'state':sheet['J'+str(point)].value,
                                            'church':sheet['K'+str(point)].value,
                                            'zone':sheet['L'+str(point)].value,
                                            }}

                    self.errors.setText("")
                    self.welcomeuser(**self.child_info[self.username.text()])
            else:
                self.errors.setText("Username does not exist")




    def openchild2(self):
        window2.setCurrentIndex(3)

    def welcomeuser(self, name, password, phone, email, teacher, classe, birth_date, acc_type, country, state, church, zone):
        winster = WelcomeUser(name, password, phone, email, teacher, classe, birth_date, acc_type, country, state, church, zone)
        window2.addWidget(winster)
        window2.setCurrentIndex(7)

class WelcomeUser(QDialog):
    def __init__(self, name, password, phone, email, teacher, classe, birth_date, acc_type, country, state, church, zone):
        super(WelcomeUser,self).__init__()

        self.name = name
        self.password = password
        self.phone = phone
        self.email = email
        self.teacher = teacher
        self.classe = classe
        self.birth_date = birth_date
        self.acc_type = acc_type
        self.country = country
        self.state = state
        self.church = church
        self.zone = zone

        loadUi("childpage3.ui", self)
        text = self.label.text()
        text = text.replace("[Child Name]", name)
        # self.label_2.setPixmap(QPixmap("curriculum.jpg"))
        # self.label_4.setPixmap(QPixmap("assignment.jpg"))
        # self.label_3.setPixmap(QPixmap("games.jpg"))

        self.label_2.setStyleSheet(self.label_2.styleSheet().replace(':/newPrefix/', ''))
        self.label_4.setStyleSheet(self.label_4.styleSheet().replace(':/newPrefix/', ''))
        self.label_3.setStyleSheet(self.label_3.styleSheet().replace(':/newPrefix/', ''))

        self.label.setStyleSheet("background-image: url(welcome.jpg); font: 16pt 'MS Shell Dlg 2'; background-color: blue; color: white;")
        self.label.setText(f"Welcome {name}")
        self.gamesfun.clicked.connect(lambda: self.loadgames(name=name))
        self.curriclum.clicked.connect(self.loadcurriculum)

    def compile_age(self, birth_date):
        listed = str(birth_date).split('/')
        print(listed)
        birth = QDate(int(listed[0]), int(listed[1]), int(listed[2]))

        birth_year = birth.year()
        birth_month = birth.month()
        birth_day = birth.day()

        current = QDate.currentDate()
        # getting year and month day of current day
        current_year = current.year()
        current_month = current.month()
        current_day = current.day()

        # coverting dates into date object
        birth_date = datetime.date(birth_year, birth_month, birth_day)
        current_date = datetime.date(current_year, current_month, current_day)

        # getting difference in both the dates
        difference = current_date - birth_date

        # getting days from the difference
        difference = difference.days
        years = difference / 365.2422
        years = round(years)
        return years
    def loadcurriculum(self):
        age = self.compile_age(str(self.birth_date).replace(' 00:00:00', '').replace('-', '/'))

        self.windows = CurriculumPage(age)
        window2.addWidget(self.windows)
        window2.setCurrentIndex(window2.currentIndex()+span)

    def loadgames(self, name):
        windows = GamesPage(name)
        window2.addWidget(windows)
        window2.setCurrentIndex(window2.currentIndex()+1)

class CurriculumPage(QDialog):
    def __init__(self, age):
        super(CurriculumPage, self).__init__()
        #loadUi("curriculum.ui", self)
        #self.label.setPixmap(QPixmap("C:\\Users\\HP\\PycharmProjects\\untitled\\GenChamps\\July\\0009"))
        self.WebBrowser = QAxContainer.QAxWidget(self)
        self.WebBrowser.setFocusPolicy(Qt.StrongFocus)
        self.WebBrowser.setControl("{8856F961-340A-11D0-A96B-00C04FD705A2}")
        self.WebBrowser.setFixedSize(354, 591)
        # convert system path to web path
        if age >= 3 and age < 6:
            print(age)
            f = Path(f"{str(os.getcwd())}\lcbcforpresc_jul_2160df2e79b992b.pdf").as_uri()
        else:
            print(age)
            f = Path(f'{str(os.getcwd())}\lcbcforupper_jul_2160df2ecaed84c.pdf').as_uri()
        # load object
        self.WebBrowser.dynamicCall('Navigate(const QString&)', f)
        self.lacks = QCommandLinkButton("Back", self)
        self.lacks.setStyleSheet('background-color: blue; color: white;')
        self.lacks.setGeometry(10, 450, 185, 41)
        self.lacks.clicked.connect(self.exits)

    def exits(self):
        window2.setCurrentIndex(window2.currentIndex()-1)

class GamesPage(QDialog):
    def __init__(self, name):
        super(GamesPage, self).__init__()
        loadUi("games.ui", self)
        self.name = name
        # self.label.setPixmap(QPixmap("word_search.jpg"))
        # self.label_2.setPixmap(QPixmap("adventure.jpg"))
        # self.label_3.setPixmap(QPixmap("puzzle.jpg"))
        # self.label_6.setPixmap(QPixmap("mutiplat.jpg"))
        # self.label_4.setPixmap(QPixmap("multiplay.jpg"))
        # self.label_5.setPixmap(QPixmap("scramble.jpg"))

        self.label.setStyleSheet(self.label.styleSheet().replace(':/newPrefix/', ''))
        self.label_2.setStyleSheet(self.label_2.styleSheet().replace(':/newPrefix/', ''))
        self.label_3.setStyleSheet(self.label_3.styleSheet().replace(':/newPrefix/', ''))
        self.label_6.setStyleSheet(self.label_6.styleSheet().replace(':/newPrefix/', ''))
        self.label_4.setStyleSheet(self.label_4.styleSheet().replace(':/newPrefix/', ''))
        self.label_5.setStyleSheet(self.label_5.styleSheet().replace(':/newPrefix/', ''))
        self.exits.clicked.connect(self.go_back)
        self.word_scramble.clicked.connect(self.play)
        self.quiz.clicked.connect(self.vic)

    def vic(self):
        today = QDate.currentDate()
        if today.day() == 1:
            from quiz_game import kbc
            self.differ_process = False
            while not self.differ_process:
                file = open("point.txt", "r")
                if file.read() == '1000000':
                    file.close()

                    file = open("point.txt", 'w')
                    file.write('0')
                    file.close()

                    conn = sqlite3.connect("children.db")
                    cursor = conn.execute("SELECT UserName, Coins from children")
                    for row in cursor:
                        if row[0] == self.name:
                            conn.execute(f"Update children set Coins = {int(row[1])+1000000} where UserName = {row[0]}")
                            conn.commit()
                            cursor.close()
                            break
                        else:
                            pass

                    w = CongratulatePoints('1000000')
                    window2.addWidget(w)
                    window2.setCurrentIndex(window2.currentIndex()+1)
                    self.differ_process = True

        else:
            QMessageBox.about(self, 'Sorry', 'This game only opens every month')

    def play(self):
        guide = Guide1()
        window2.addWidget(guide)
        window2.setCurrentIndex(window2.currentIndex()+1)

    def go_back(self):
        window2.setCurrentIndex(window2.currentIndex()-1)
class CongratulatePoints(QDialog):
    def __init__(self, points):
        super(CongratulatePoints, self).__init__()
        self.pointed = points
        loadUi("endings.ui", self)
        self.label_7.setText(self.label_7.text().replace('kkp', self.pointed))
        self.label.setStyleSheet("background-image: url(coin.jpg)")
class Guide1(QDialog):
    def __init__(self):
        super(Guide1, self).__init__()
        loadUi("guide_1.ui", self)
        self.starter.clicked.connect(self.scramble)
    def scramble(self):
        we = WordScramblers()
        window2.addWidget(we)
        window2.setCurrentIndex(window2.currentIndex()+1)

class WordScramblers(QDialog):
    def __init__(self):
        super(WordScramblers, self).__init__()
        loadUi("word_scramblers.ui", self)
        self.widget.show()
        self.label_5.setStyleSheet("background-image: url(coin.jpg)")
        self.load_game()

    def load_game(self):
        self.words = ["Jesus", 'Heaven', 'God', 'Jerusalem', 'Israel', 'Judas', 'Judea', 'Egypt', 'Moses', 'Abraham',
                 'Jacob',
                 'Kingdom', 'Angel', 'Lamb', 'Faith', 'Knowledge', 'Wisdom', 'Grace', 'Peace', 'Joy', 'Hope',
                 'Kindness',
                 'Patuence', 'Contentment', 'Satisfaction', 'Grace', 'Happiness', 'Excitement', 'Longsuffering',
                 'Isaiah',
                 'Mary', 'Joseph', 'Cain', 'Abel', 'Adam', 'Eve', 'Love', 'Champion', 'Bible', 'Daniel', 'Christmas', 'New', 'Year',
                 'Virtue', 'Strength', 'Victory']
        self.word = random.randint(0, len(self.words) - 1)
        self.num = self.words[self.word].upper()
        self.s2 = ''.join(random.sample(self.num, len(self.num)))
        self.label_3.setText(self.label_3.text().replace('tEXT', str(self.s2)))
        self.submit.clicked.connect(self.check)
        self.end.clicked.connect(self.calc)

    def calc(self):
        self.widget.hide()
        loadUi("second_ending.ui", self)
        self.widget_2.show()
        self.label_7.setText(self.label_7.text().replace('kkp', self.coin.text()))
        self.label.setStyleSheet("background-image: url(coin.jpg)")
        self.pushButton.clicked.connect(self.play_again)
        self.commandLinkButton.clicked.connect(self.go_back)

    def go_back(self):
        window2.setCurrentIndex(window2.currentIndex()-1)

    def play_again(self):
        self.widget.show()
        self.widget_2.hide()

    def check(self):
        if self.answer.text().lower().replace(' ', '') == self.num.lower():
            engine = pyttsx3.init()
            voices = engine.getProperty('voices')
            engine.setProperty('voice', voices[0].id)

            engine.say(f'You have gotten the word {self.num} Correctly!')
            engine.runAndWait()

            #QMessageBox.about(self, 'Congratulations', f'You have gotten the word {self.num} correctly')
            self.coin.setText(str(int(self.coin.text())+10))
        else:
            engine = pyttsx3.init()
            voices = engine.getProperty('voices')
            engine.setProperty('voice', voices[0].id)

            engine.say(f'Sorry, you have spelt the word {self.num} Incorrectly!')
            engine.runAndWait()

        self.word = random.randint(0, len(self.words) - 1)
        self.num = self.words[self.word].upper()
        self.s = ''.join(random.sample(self.num, len(self.num)))
        self.label_3.setText(self.s)

    def resume(self):
        self.load_game()


class ChildPage2(QDialog):
    def __init__(self):
        super(ChildPage2, self).__init__()
        loadUi("childpage2.ui", self)
        self.label_7.setPixmap(QPixmap("sunday.jpg"))

        file = xl.load_workbook("Country-codes.xlsx")
        sheet = file['Sheet1']

        self.file2 = xl.load_workbook("states.xlsx")
        self.sheet2 = self.file2['Sheet1']

        self.countries = {}
        for i in range(244):
            self.countries[sheet['A'+str(i+1)].value] = sheet['B'+str(i+1)].value
            self.country.addItem(sheet['A'+str(i+1)].value)

        for i in range(41002):
            if str(self.sheet2['B'+str(i+1)].value).lower() in str(self.country.currentText()).lower():
                sets = str(self.sheet2['B'+str(i+1)].value)
                for i in range(41002):
                    if str(self.sheet2['B'+str(i+1)].value) == sets:
                        self.states.addItem(str(self.sheet2['A'+str(i)].value))
                break
        self.country.currentTextChanged.connect(self.always_change)

    def always_change(self):
        self.states.clear()
        for i in range(41002):
            if str(self.sheet2['B'+str(i+1)].value).lower() in str(self.country.currentText()).lower():
                sets = str(self.sheet2['B'+str(i+1)].value)
                for i in range(41002):
                    if str(self.sheet2['B'+str(i+1)].value) == sets:
                        self.states.addItem(str(self.sheet2['A'+str(i+1)].value))
                break

        self.country.currentTextChanged.connect(self.change)
        self.phone_no.setText(f"({self.countries['Afghanistan ']})")
        self.phone.setPlaceholderText(f"Phone Number")
        self.submitter.clicked.connect(self.confirm_all)
        self.phone.setValidator(QIntValidator())
        self.phone.textChanged.connect(self.collect)

    def collect(self):
        try:
            if str(self.phone.text()[0]) == "0":
                text = self.phone.text()
                text = text.replace('0', '')
                self.phone.setText(text)
        except:
            pass

    def change(self):
        self.phone_no.setText(f"({self.countries[str(self.country.currentText())]})")
    def confirm_all(self):
        self.fullnames = self.fullname.text()
        self.emails = self.email.text()
        self.passwords = self.password.text()
        self.confirms_pass = self.confirm_pass.text()
        self.countrsy = self.country.currentText()
        self.anti_deep_states = self.states.currentText()
        self.phones = self.phone.text()
        self.parent_names = self.parent_name.text()

        if self.fullnames == '':
            self.errors.setText("Please fill out the fullname field")
        else:
            if self.emails == '':
                self.errors.setText("Please fill out the email field")
            else:
                if len(self.passwords) < 6:
                    self.errors.setText("Your password must not be less than 6 characters")
                else:
                    if self.confirms_pass != self.password:
                        self.errors.setText("Please the password you typed is not the same as the one\nthat you typed to confirm.")
                    else:
                        if len(self.phones) != 10:
                            self.errors.setText("Please enter a valid phone number")
                        else:
                            op = 2
                            conn = sqlite3.connect("children.db")
                            cursor = conn.execute(f""""INSERT INTO children
                                                      (UserName, Password, PhoneNumber, Email, Teacher, Class, BirthDate, AccountType, Country, State, Church, Zone, ParentName, Coins)
                                                      VALUES
                                                      ('{self.fullnames}', '{self.passwords}', '{self.phones}', '{self.emails}', null, null, {f"{self.dateEdit.date().year()}/{self.dateEdit.date().month()}/{self.dateEdit.date().day()}"}, "Outreach Account", '{self.countrsy}', '{self.anti_deep_states}', null, null, '{self.parent_names}')""")
                            conn.commit()
                            cursor.close()


app = QApplication(sys.argv)

class First(QDialog):
    def __init__(self):
        super(First, self).__init__()
        self.label = QLabel(self)
        self.label.setGeometry(0, 0, 357, 599)
        self.label.setPixmap(QPixmap("picture.jpg"))
        bar = QProgressBar(self)
        bar.setGeometry(30, 400, bar.width()+100, bar.height())
        bar.setValue(0)
        self.lf = 0
        for i in range(100):
            i2 = 3000 #random.randint(2000, 5000)
            QTimer.singleShot(i2, lambda: bar.setValue(bar.value()+1))
            self.lf += i2


def new():
    window2.addWidget(window)
    window2.setCurrentIndex(1)
span = 1
window = Window()
window3 = ChildPage()
window4 = TeacherPage()
window5 = ChildPage2()
window6 = Guide1()
window7 = Curriculums()
window8 = AssociateTeacher("", "", "", "", "")
window2 = QStackedWidget()

window2.setGeometry(100, 100, 357, 599)
window2.addWidget(window) #0
window2.addWidget(window3) #1
window2.addWidget(window4) #2
window2.addWidget(window5)
window2.addWidget(window6)
window2.addWidget(window7)
window2.addWidget(window8)
window2.setWindowTitle("The LoveWorld Children Curriculum App")
window2.setMaximumSize(357, 599)
window2.setMinimumSize(357, 599)
window2.show()
app.exec_()
